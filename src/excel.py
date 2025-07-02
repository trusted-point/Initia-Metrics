import asyncio
import openpyxl
import emoji

from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter

from sys import exit
from utils.logger import log
from src.aio_calls import AioHttpCalls
from src.decoder import KeysUtils
from src.mongodb import MongoDBHandler

class Excel:
    def __init__(
        self,
        aio_session: AioHttpCalls,
        mongo: MongoDBHandler,

        excel_file_name: str,
        excel_sheet_name: str
    ):
        self.headers = []
        self.aio_session = aio_session
        self.mongo = mongo

        self.excel_file_name = excel_file_name
        self.excel_sheet_name = excel_sheet_name

        # STYLES
        self.bold_border = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000")
        )
        self.light_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        self.header_font = Font(bold=True, size=12)
        self.range_font = Font(bold=False, size=10, color='707070')
        self.balck_bold_font = Font(color="000000", bold=True)

        self.light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        self.light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        self.light_yellow_fill = PatternFill(start_color="FFFFC5", end_color="FFFFC5", fill_type="solid")
        self.light_pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        #FFC0CB

    async def get_validators(self):
        validators = []
        next_key = None

        async def fetch_with_retry(next_key, retries=3):
            for attempt in range(retries):
                try:
                    result = await self.aio_session.fetch_validators(
                        status=None,
                        pagination_limit=100,
                        next_key=next_key
                    )
                    if result and 'validators' in result:
                        if attempt > 0:
                            log.info(f"Successfully fetched {len(result['validators'])} validators after {attempt + 1} attempt(s).")
                        return result
                    else:
                        raise ValueError("Invalid response")
                    
                except Exception as e:
                    if attempt < retries - 1:
                        log.warning(f"Retrying block validators request (attempt {attempt + 1}) due to: {e}")
                        await asyncio.sleep(3)
                    else:
                        log.error(f"Failed to fetch validators after {retries} attempt(s).")
                        return

        while True:
            page = await fetch_with_retry(next_key)
            if not page:
                exit(5)
            validators.extend(page["validators"])
            next_key = page.get("pagination", {}).get("next_key")
            if not next_key:
                break

        validators_result = []
        if validators:
            for validator in validators:
                _consensus_pub_key = validator["consensus_pubkey"]["key"]
                _moniker = validator["description"]["moniker"]
                _valoper = validator["operator_address"]
                _hex = KeysUtils.pub_key_to_consensus_hex(pub_key=_consensus_pub_key)
                _valcons = KeysUtils.pub_key_to_bech32(pub_key=_consensus_pub_key, address_refix="valcons")
                _wallet = KeysUtils.valoper_to_account(valoper=_valoper)
                _tombstoned = await self.aio_session.get_validator_tomb(valcons=_valcons)
                if _tombstoned is None:
                    _tombstoned = False

                validators_result.append({
                    "moniker": _moniker,
                    "hex": _hex,
                    "valoper": _valoper,
                    "consensus_pubkey": _consensus_pub_key,
                    "valcons": _valcons,
                    "wallet": _wallet,
                    "tombstoned": _tombstoned
                })

        log.info(f"Succesfully fecthed {len(validators_result)} validators")
        return validators_result
    
    def create_excel_sheet(self) -> tuple[Worksheet, Workbook]:
        try:
            workbook = openpyxl.load_workbook(self.excel_file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook[f"Sheet"])
        if f"{self.excel_sheet_name}" in workbook.sheetnames:
            workbook.remove(workbook[f"{self.excel_sheet_name}"])
        sheet = workbook.create_sheet(title=f"{self.excel_sheet_name}")
        return sheet, workbook

    def create_main_table(
        self,
        sheet: Worksheet,
        workbook: Workbook,
        validators: list[dict],
        days_stats: list[dict],
        slashes: list[dict],
        proposals: list[dict]
    ) -> None:
        
        self.headers = ["Moniker", "Validator Address", "Tombstoned", f"Voted Proposals ({len(proposals)})", "Jails", "Signed ✔", "Missed ❌", "Proposed 🚩", "Uptime"]
        
        sorted_dates = sorted(days_stats, key=lambda entry: entry['date_start_height'])

        for validator in validators:
            total_signed_blocks = 0
            total_missed_blocks = 0
            total_proposed_blocks = 0
            _hex = validator['hex']
            _valoper = validator['valoper']

            for day in sorted_dates:
                day_signed_blocks = day['validators'].get(_hex, {}).get('signed_blocks', 0)
                day_missed_blocks = day['validators'].get(_hex, {}).get('missed_blocks', 0)
                day_proposed_blocks = day['validators'].get(_hex, {}).get('proposed_blocks', 0)
                total_signed_blocks += day_signed_blocks
                total_missed_blocks += day_missed_blocks
                total_proposed_blocks += day_proposed_blocks

            validator['total_signed_blocks'] = total_signed_blocks
            validator['total_missed_blocks'] = total_missed_blocks
            validator['total_proposed_blocks'] = total_proposed_blocks

            # Slashes
            validator['total_slashes'] = 0
            validator['slashes'] = []
            for entry in slashes:
                if entry['_id'] == _valoper and entry['slashes']:
                    validator['total_slashes'] = len(entry['slashes'])
                    validator['slashes'] = entry['slashes']
                    break
            # Gov
            total_voted_proposals = 0
            for proposal in proposals:
                vote = proposal['validators'].get(_valoper,{}).get('tx_hash')
                if vote:
                    total_voted_proposals += 1
            validator['total_voted_proposals'] = total_voted_proposals

        sorted_validators = sorted(validators, key=lambda x: x["total_signed_blocks"] + x["total_missed_blocks"], reverse=True)

        for col_num, header in enumerate(self.headers, start=1):
            header_cell = sheet.cell(row=1, column=col_num, value=header)
            header_cell.font = self.header_font
            header_cell.alignment = self.center_alignment
            header_cell.fill = self.header_fill
            header_cell.border = self.bold_border

        day_to_col_num = {}
        col_num = len(self.headers) + 1
        for day in sorted_dates:
            day_cell = sheet.cell(row=1, column=col_num, value=day['_id'])
            day_cell.font = self.header_font
            day_cell.alignment = self.center_alignment
            day_cell.fill = self.header_fill
            day_cell.border = self.bold_border
            day_to_col_num[day['_id']] = col_num
            col_num += 1

        row_num = 2

        for validator in sorted_validators:
            moniker = emoji.replace_emoji(validator["moniker"], replace='')

            valoper = validator["valoper"]
            hex_addr = validator["hex"]
            tombstoned = str(validator["tombstoned"])
            gov_participation =  validator['total_voted_proposals'] / len(proposals)
            total_signed = validator["total_signed_blocks"]
            total_missed = validator["total_missed_blocks"]
            total_proposed = validator["total_proposed_blocks"]
            total_uptime = round((
                (total_signed / (total_signed + total_missed))
                if (total_signed + total_missed) > 0 else 0.0
            ),6)

            uptime_days = []
            for day in sorted_dates:
                signed = day['validators'].get(hex_addr, {}).get('signed_blocks', 0)
                missed = day['validators'].get(hex_addr, {}).get('missed_blocks', 0)
                day_uptime = round((
                    (signed / (signed + missed))
                    if (signed + missed) > 0 else 0.0
                ),6)
                
                if (signed + missed) > 0:
                    uptime_days.append(day_uptime)

                    if day_uptime >= 0.90:
                        uptime_color = self.light_green_fill
                    elif day_uptime >= 0.60:
                        uptime_color = self.light_yellow_fill
                    else:
                        uptime_color = self.light_red_fill

                    uptime_cell = sheet.cell(row=row_num, column=day_to_col_num[day['_id']], value=day_uptime)
                    uptime_cell.font = self.balck_bold_font
                    uptime_cell.alignment = self.center_alignment

                    uptime_cell.fill = uptime_color
                    uptime_cell.border = self.light_border
                    uptime_cell.number_format = '0.0000%'

            cell_moniker = sheet.cell(row=row_num, column=1, value=moniker)
            cell_moniker.alignment = self.left_alignment
            cell_moniker.border = self.bold_border

            cell_valoper = sheet.cell(row=row_num, column=2, value=valoper)
            cell_valoper.alignment = self.left_alignment
            cell_valoper.border = self.bold_border

            cell_tombstoned = sheet.cell(row=row_num, column=3, value=tombstoned)
            false_font = Font(bold=True, size=12, color='008000')
            true_font = Font(bold=True, size=12, color='FF0000')
            cell_tombstoned.font = true_font if tombstoned == 'True' else false_font
            cell_tombstoned.alignment = self.center_alignment
            cell_tombstoned.border = self.bold_border
            
            if gov_participation >= 0.70:
                gov_font = Font(bold=True, size=12, color='008000')

            elif gov_participation >= 0.30:
                gov_font = Font(bold=True, size=12, color='FFA500')
            else:
                gov_font = Font(bold=True, size=12, color='FF0000')

            cell_gov = sheet.cell(row=row_num, column=4, value=validator['total_voted_proposals'])
            cell_gov.font = gov_font
            cell_gov.alignment = self.center_alignment
            cell_gov.border = self.bold_border

            cell_slashes = sheet.cell(row=row_num, column=5, value=validator['total_slashes'])
            cell_slashes.font = true_font if validator['total_slashes'] else false_font
            cell_slashes.alignment = self.center_alignment
            cell_slashes.border = self.bold_border

            cell_total_signed = sheet.cell(row=row_num, column=6, value=total_signed)
            cell_total_signed.alignment = self.center_alignment
            cell_total_signed.border = self.bold_border

            cell_total_missed = sheet.cell(row=row_num, column=7, value=total_missed)
            cell_total_missed.alignment = self.center_alignment
            cell_total_missed.border = self.bold_border

            cell_total_proposed = sheet.cell(row=row_num, column=8, value=total_proposed)
            cell_total_proposed.alignment = self.center_alignment
            cell_total_proposed.border = self.bold_border

            if total_uptime >= 0.90:
                uptime_font = Font(bold=True, size=12, color='008000')

            elif total_uptime >= 0.60:
                uptime_font = Font(bold=True, size=12, color='FFA500')
            else:
                uptime_font = Font(bold=True, size=12, color='FF0000')

            cell_total_uptime = sheet.cell(row=row_num, column=9, value=total_uptime)
            cell_total_uptime.font = uptime_font
            cell_total_uptime.alignment = self.center_alignment
            cell_total_uptime.border = self.bold_border
            cell_total_uptime.number_format = '0.0000%'

            row_num += 1

        min_width = 10
        padding = 3
        for col_cells in sheet.columns:

            lengths = [
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            ]
            max_length = max(lengths)
            col_letter = get_column_letter(col_cells[0].column)
            sheet.column_dimensions[col_letter].width = max(max_length, min_width) + padding
        
        workbook.save(self.excel_file_name)
        log.info(f"Excel file updated: {self.excel_file_name} | {self.excel_sheet_name}")

    def create_governance_table(
        self,
        sheet: Worksheet,
        workbook: Workbook,
        validators: list[dict],
        proposals: list[dict]
    ) -> None:
        
        self.headers = ["Moniker", "Validator Address", f"Voted Proposals ({len(proposals)})", "Participation"]
        

        for validator in validators:
            _valoper = validator['valoper']
            total_voted_proposals = 0
            for proposal in proposals:
                vote = proposal['validators'].get(_valoper,{}).get('tx_hash')
                if vote:
                    total_voted_proposals += 1
            validator['total_voted_proposals'] = total_voted_proposals
        sorted_validators = sorted(validators, key=lambda x: x["total_voted_proposals"], reverse=True)
        sorted_proposals = sorted(proposals, key=lambda x: int(x["_id"]), reverse=False)

        proposal_to_col_num = {}
        col_num = len(self.headers) + 1
        for proposal in sorted_proposals:
            prop_cell = sheet.cell(row=1, column=col_num, value=int(proposal['_id']))
            prop_cell.font = self.header_font
            prop_cell.alignment = self.center_alignment
            prop_cell.fill = self.header_fill
            prop_cell.border = self.bold_border
            proposal_to_col_num[proposal['_id']] = col_num
            col_num += 1

        for col_num, header in enumerate(self.headers, start=1):
            header_cell = sheet.cell(row=1, column=col_num, value=header)
            header_cell.font = self.header_font
            header_cell.alignment = self.center_alignment
            header_cell.fill = self.header_fill
            header_cell.border = self.bold_border

        row_num = 2

        for validator in sorted_validators:
            moniker = emoji.replace_emoji(validator["moniker"], replace='')
            valoper = validator["valoper"]
            gov_participation =  round(validator['total_voted_proposals'] / len(proposals),6)

            for proposal in sorted_proposals:
                vote = None
                if proposal['validators'].get(valoper):
                    vote = proposal['validators'][valoper]["option"]

                if vote is None:
                    vote_color = None
                elif vote == "Yes":
                    vote_color = self.light_green_fill
                elif vote == "No":
                    vote_color = self.light_pink_fill
                elif vote == "NoWithVeto":
                    vote_color = self.light_red_fill
                elif vote == "Abstain":
                    vote_color = self.light_yellow_fill
                else:
                    vote_color = self.light_yellow_fill

                if vote_color:
                    vote_cell = sheet.cell(row=row_num, column=proposal_to_col_num[proposal['_id']], value=vote)
                    vote_cell.font = self.balck_bold_font
                    vote_cell.alignment = self.center_alignment
                    vote_cell.border = self.light_border
                    vote_cell.fill = vote_color

            cell_moniker = sheet.cell(row=row_num, column=1, value=moniker)
            cell_moniker.alignment = self.left_alignment
            cell_moniker.border = self.bold_border

            cell_valoper = sheet.cell(row=row_num, column=2, value=valoper)
            cell_valoper.alignment = self.left_alignment
            cell_valoper.border = self.bold_border
            
            if gov_participation >= 0.70:
                gov_font = Font(bold=True, size=12, color='008000')

            elif gov_participation >= 0.30:
                gov_font = Font(bold=True, size=12, color='FFA500')
            else:
                gov_font = Font(bold=True, size=12, color='FF0000')

            cell_gov = sheet.cell(row=row_num, column=3, value=validator['total_voted_proposals'])
            cell_gov.font = gov_font
            cell_gov.alignment = self.center_alignment
            cell_gov.border = self.bold_border

            cell_gov_participation = sheet.cell(row=row_num, column=4, value=gov_participation)
            cell_gov_participation.font = gov_font
            cell_gov_participation.alignment = self.center_alignment
            cell_gov_participation.border = self.bold_border
            cell_gov_participation.number_format = '0.0000%'
            row_num += 1

        min_width = 10
        padding = 3
        for col_cells in sheet.columns:

            lengths = [
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            ]
            max_length = max(lengths)
            col_letter = get_column_letter(col_cells[0].column)
            sheet.column_dimensions[col_letter].width = max(max_length, min_width) + padding
        
        workbook.save(self.excel_file_name)
        log.info(f"Excel file updated: {self.excel_file_name} | {self.excel_sheet_name}")

    def create_oracle_table(
        self,
        sheet: Worksheet,
        workbook: Workbook,
        validators: list[dict],
        days_stats: list[dict],
    ) -> None:
        
        self.headers = ["Moniker", "Validator Address", "Prices ✔", "No Prices ❌", "Uptime"]
        
        sorted_dates = sorted(days_stats, key=lambda entry: entry['date_start_height'])

        for validator in validators:
            total_signed_oracle = 0
            total_missed_oracle = 0
            _hex = validator['hex']
            for day in sorted_dates:
                day_signed_oracle = day['validators'].get(_hex, {}).get('signed_oracle', 0)
                day_missed_oracle = day['validators'].get(_hex, {}).get('missed_oracle', 0)
                total_signed_oracle += day_signed_oracle
                total_missed_oracle += day_missed_oracle
            validator['total_signed_oracle'] = total_signed_oracle
            validator['total_missed_oracle'] = total_missed_oracle

        sorted_validators = sorted(validators, key=lambda x: x["total_signed_oracle"] + x["total_missed_oracle"], reverse=True)

        for col_num, header in enumerate(self.headers, start=1):
            header_cell = sheet.cell(row=1, column=col_num, value=header)
            header_cell.font = self.header_font
            header_cell.alignment = self.center_alignment
            header_cell.fill = self.header_fill
            header_cell.border = self.bold_border

        day_to_col_num = {}
        col_num = len(self.headers) + 1
        for day in sorted_dates:
            day_cell = sheet.cell(row=1, column=col_num, value=day['_id'])
            day_cell.font = self.header_font
            day_cell.alignment = self.center_alignment
            day_cell.fill = self.header_fill
            day_cell.border = self.bold_border
            day_to_col_num[day['_id']] = col_num
            col_num += 1

        row_num = 2

        for validator in sorted_validators:
            moniker = emoji.replace_emoji(validator["moniker"], replace='')

            valoper = validator["valoper"]
            hex_addr = validator["hex"]
            total_signed_oracle = validator["total_signed_oracle"]
            total_missed_oracle = validator["total_missed_oracle"]
            total_uptime = round((
                (total_signed_oracle / (total_signed_oracle + total_missed_oracle))
                if (total_signed_oracle + total_missed_oracle) > 0 else 0.0
            ),6)

            uptime_days = []
            for day in sorted_dates:
                signed = day['validators'].get(hex_addr, {}).get('signed_oracle', 0)
                missed = day['validators'].get(hex_addr, {}).get('missed_oracle', 0)
                day_uptime = round((
                    (signed / (signed + missed))
                    if (signed + missed) > 0 else 0.0
                ),6)
                
                if (signed + missed) > 0:
                    uptime_days.append(day_uptime)

                    if day_uptime >= 0.90:
                        uptime_color = self.light_green_fill
                    elif day_uptime >= 0.60:
                        uptime_color = self.light_yellow_fill
                    else:
                        uptime_color = self.light_red_fill

                    uptime_cell = sheet.cell(row=row_num, column=day_to_col_num[day['_id']], value=day_uptime)
                    uptime_cell.font = self.balck_bold_font
                    uptime_cell.alignment = self.center_alignment

                    uptime_cell.fill = uptime_color
                    uptime_cell.border = self.light_border
                    uptime_cell.number_format = '0.0000%'

            cell_moniker = sheet.cell(row=row_num, column=1, value=moniker)
            cell_moniker.alignment = self.left_alignment
            cell_moniker.border = self.bold_border

            cell_valoper = sheet.cell(row=row_num, column=2, value=valoper)
            cell_valoper.alignment = self.left_alignment
            cell_valoper.border = self.bold_border

            cell_total_signed = sheet.cell(row=row_num, column=3, value=total_signed_oracle)
            cell_total_signed.alignment = self.center_alignment
            cell_total_signed.border = self.bold_border

            cell_total_missed = sheet.cell(row=row_num, column=4, value=total_missed_oracle)
            cell_total_missed.alignment = self.center_alignment
            cell_total_missed.border = self.bold_border

            if total_uptime >= 0.90:
                uptime_font = Font(bold=True, size=12, color='008000')

            elif total_uptime >= 0.60:
                uptime_font = Font(bold=True, size=12, color='FFA500')
            else:
                uptime_font = Font(bold=True, size=12, color='FF0000')

            cell_total_uptime = sheet.cell(row=row_num, column=5, value=total_uptime)
            cell_total_uptime.font = uptime_font
            cell_total_uptime.alignment = self.center_alignment
            cell_total_uptime.border = self.bold_border
            cell_total_uptime.number_format = '0.0000%'

            row_num += 1

        min_width = 10
        padding = 3
        for col_cells in sheet.columns:

            lengths = [
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            ]
            max_length = max(lengths)
            col_letter = get_column_letter(col_cells[0].column)
            sheet.column_dimensions[col_letter].width = max(max_length, min_width) + padding
        
        workbook.save(self.excel_file_name)
        log.info(f"Excel file updated: {self.excel_file_name} | {self.excel_sheet_name}")

    def create_slashes_table(
        self,
        sheet: Worksheet,
        workbook: Workbook,
        validators: list[dict],
        slashes: list[dict],
    ) -> None:
        
        slashes_sorted = sorted(slashes, key=lambda e: len(e["slashes"]), reverse=True)
        
        self.headers = ["Moniker", "Validator Address", "Slashing Height", "Slashing Date", "Slashing Time (UTC)", "Total Jails"]
        for col_num, header in enumerate(self.headers, start=1):
            header_cell = sheet.cell(row=1, column=col_num, value=header)
            header_cell.font = self.header_font
            header_cell.alignment = self.center_alignment
            header_cell.fill = self.header_fill
            header_cell.border = self.bold_border

        row_num = 2

        for entry in slashes_sorted:
            valoper       = entry["_id"]
            moniker       = next(
                (emoji.replace_emoji(v["moniker"], replace='')
                for v in validators if v["valoper"] == valoper),
                "N/A"
            )
            slashes_list  = entry["slashes"]
            count         = len(slashes_list)
            start_row     = row_num

            for slash in slashes_list:
                height = slash["height"]
                date   = slash["date"]
                day    = date.split("T")[0]
                time   = date.split("T")[1].split(".")[0]

                sheet.cell(row=row_num, column=3, value=height).alignment = self.center_alignment
                sheet.cell(row=row_num, column=4, value=day).alignment    = self.center_alignment
                sheet.cell(row=row_num, column=5, value=time).alignment   = self.center_alignment
                row_num += 1

            end_row = row_num - 1

            if count > 1:
                for col_letter, value in (("A", moniker), ("B", valoper), ("F", count)):
                    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
                    sheet.merge_cells(rng)
                    cell = sheet[f"{col_letter}{start_row}"]
                    cell.value     = value
                    cell.alignment = self.left_alignment if col_letter in ("A","B") else self.center_alignment

            else:
                sheet.cell(row=start_row, column=1, value=moniker).alignment = self.left_alignment
                sheet.cell(row=start_row, column=2, value=valoper).alignment = self.left_alignment
                cell = sheet.cell(row=start_row, column=6, value=count)
                cell.alignment = self.center_alignment

        min_width = 10
        padding = 3
        for col_cells in sheet.columns:

            lengths = [
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            ]
            max_length = max(lengths)
            col_letter = get_column_letter(col_cells[0].column)
            sheet.column_dimensions[col_letter].width = max(max_length, min_width) + padding
        
        workbook.save(self.excel_file_name)
        log.info(f"Excel file updated: {self.excel_file_name} | {self.excel_sheet_name}")

    async def start(self):
        log.info("Fetching validators…")
        validators = await self.get_validators()
        if not validators:
            log.error("No validators found, exiting.")
            exit(5)

        log.info("Fetching daily stats…")
        days_stats = await self.mongo.get_validator_stats_days()
        if not days_stats:
            log.error("No daily stats found, exiting.")
            exit(5)

        log.info("Fetching slashes…")
        slashes = await self.mongo.get_slashes()
        if not slashes:
            log.error("No slashes found, exiting.")
            exit(5)

        log.info("Fetching proposals…")
        proposals = await self.mongo.get_processed_proposals()
        if not proposals:
            log.error("No proposals found, exiting.")
            exit(5)

        log.info("All data fetched successfully. Continuing…")

        sheet, workbook = self.create_excel_sheet()

        if self.excel_sheet_name == "Main":
            self.create_main_table(
                sheet=sheet,
                workbook=workbook,
                validators=validators,
                days_stats=days_stats,
                slashes=slashes,
                proposals=proposals
            )

        elif self.excel_sheet_name == "Gov":
            self.create_governance_table(
                sheet=sheet,
                workbook=workbook,
                validators=validators,
                proposals=proposals
            )
        
        elif self.excel_sheet_name == "Oracle":
            self.create_oracle_table(
                sheet=sheet,
                workbook=workbook,
                validators=validators,
                days_stats=days_stats,
            )
        
        if self.excel_sheet_name == "Slashes":
            self.create_slashes_table(
                sheet=sheet,
                workbook=workbook,
                validators=validators,
                slashes=slashes,
            )






